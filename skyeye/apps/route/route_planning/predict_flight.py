# coding=utf-8

import random
import os, sys
import xlrd
import math
from osgeo import gdal, osr, ogr
import numpy as np
import django
from pyproj import Transformer

django.setup()
from django.conf import settings
from openpyxl import load_workbook

WORK_DIR = settings.BASE_DIR
from .create_fishnet_by_gdal import get_finish_net_point


class GA(object):
    def __init__(self, num_city, num_total, iteration, data, uuid_value):
        self.num_city = num_city  # 城市数量
        self.num_total = num_total  # num_total是一代种群的个体数量
        self.scores = []
        self.length_total = []
        self.uuid_value = uuid_value
        self.output = []
        self.iteration = iteration  # 遗传代数
        self.location = data
        self.ga_choose_ratio = 0.2  # 遗传概率
        self.mutate_ratio = 0.05  # 变异概率
        # fruits中存每一个个体是下标的list
        self.dis_mat = self.compute_dis_mat(num_city, data)  # 计算两点间的距离
        self.fruits = self.greedy_init(self.dis_mat, num_total, num_city)  # 尝试从每一起点开始遍历所有点的距离
        # 显示初始化后的最佳路径
        scores = self.compute_adp(self.fruits, self.length_total)  # 得到适应度数组
        sort_index = np.argsort(-scores)  # 降序排列优秀个体，这里排列的是适应度
        init_best = self.fruits[sort_index[0]]  # 最优个体
        init_best = self.location[init_best]

        # 存储每个iteration的结果，画出收敛图
        self.iter_x = [0]
        self.iter_y = [1. / scores[sort_index[0]]]  # 显示每一代最优个体的适应度

    def random_init(self, num_total, num_city):
        tmp = [x for x in range(num_city)]
        result = []
        for i in range(num_total):
            random.shuffle(tmp)  # 将序列的所有元素随机排序
            result.append(tmp.copy())
        return result

    def greedy_init(self, dis_mat, num_total, num_city):  # 从0点开始找一条最短路径 （什么方法，从当前点到下一最短点的路径）
        start_point = 0
        end_point = num_city - 1
        start_index = 1
        result = []
        for i in range(num_total):
            if num_city <= 2:
                result.append([start_point, end_point])
                continue
            result_one = [start_point]
            rest = [x for x in range(1, num_city)]  # 复制父代除起点外所有个体
            rest.remove(end_point)
            if start_index >= num_city - 1:
                if result:
                    result.append(result[i % len(result)].copy())
                else:
                    middle_points = list(range(1, num_city - 1))
                    result.append([start_point] + middle_points + [end_point])
                continue
            current = start_index
            result_one.append(current)
            rest.remove(current)  # rest表示所有城市点，现去掉起始点
            # 找到一条最近邻路径
            while len(rest) != 0:  # 从当前点到剩下所有点中距离最近的点
                tmp_min = math.inf
                tmp_choose = -1
                for x in rest:  # 找出rest中最短距离
                    if dis_mat[current][x] < tmp_min:
                        tmp_min = dis_mat[current][x]
                        tmp_choose = x

                current = tmp_choose
                result_one.append(tmp_choose)  # 一条染色体上加上一个碱基
                rest.remove(tmp_choose)
            result_one.append(end_point)
            start_index += 1  # 下一起始点开始
            result.append(result_one)  # 这里得到的是一条链
        return result

    # 计算不同城市之间的距离
    def compute_dis_mat(self, num_city, location):
        dis_mat = np.zeros((num_city, num_city))
        for i in range(num_city):
            for j in range(num_city):
                if i == j:
                    dis_mat[i][j] = np.inf
                    continue
                a = location[i]  # 这里是一个点
                b = location[j]
                tmp = np.sqrt(sum([(x[0] - x[1]) ** 2 for x in zip(a, b)]))  # 用了一个迭代器计算两点距离
                dis_mat[i][j] = tmp
        # 手动检查距离矩阵
        return dis_mat

    # 计算路径长度（闭合）
    def compute_pathlen(self, path, dis_mat):  # path为一条染色体长度N，dis_mat也为N*N维度
        try:
            a = path[0]
            b = path[1]
        except:
            import pdb
            pdb.set_trace()
        result = 0
        for i in range(len(path) - 1):
            a = path[i]
            b = path[i + 1]
            result += dis_mat[a][b]
        return result

    # 计算种群每个个体适应度
    def compute_adp(self, fruits, length_total):
        adp = []
        for fruit in fruits:
            if isinstance(fruit, int):
                import pdb
                pdb.set_trace()
            length = self.compute_pathlen(fruit, self.dis_mat)
            adp.append(1.0 / length)
            length_total.append(length)

        return np.array(adp)

    def swap_part(self, list1, list2):
        index = len(list1)
        list = list1 + list2
        list = list[::-1]
        return list[:index], list[index:]

    def ga_cross_beifen(self, x, y):
        len_ = len(x)
        assert len(x) == len(y)
        path_list = [t for t in range(len_)]  # 生成与父代个体相同长度的链
        order = list(random.sample(path_list, 2))  # path_list列表中任意取出两个，组成一个新的列表
        order.sort()
        start, end = order
        # 找到冲突点并存下他们的下标,x中存储的是y中的下标,y中存储x与它冲突的下标
        tmp = x[start:end]
        x_conflict_index = []
        for sub in tmp:
            index = y.index(sub)
            if not (index >= start and index < end):
                x_conflict_index.append(index)  # 这里的index是y片段中的位置

        y_confict_index = []
        tmp = y[start:end]
        for sub in tmp:
            index = x.index(sub)
            if not (index >= start and index < end):
                y_confict_index.append(index)

        assert len(x_conflict_index) == len(y_confict_index)

        # 交叉
        tmp = x[start:end].copy()
        x[start:end] = y[start:end]
        y[start:end] = tmp

        # 解决冲突，这种解法直接在交换区间外对换丢失的基因，不更改交换的基因片段
        for index in range(len(x_conflict_index)):
            i = x_conflict_index[index]
            j = y_confict_index[index]
            y[i], x[j] = x[j], y[i]

        assert len(set(x)) == len_ and len(set(y)) == len_
        return list(x), list(y)

    def ga_cross(self, x, y):
        len_ = len(x)
        assert x[0] == 0 and x[-1] == len_ - 1  # 验证起点终点
        assert y[0] == 0 and y[-1] == len_ - 1

        # 只在中间部分(1~len_-2)选择交叉点
        crossover_points = sorted(random.sample(range(1, len_ - 1), 2))
        start, end = crossover_points

        # 原交叉逻辑（保持不变）
        tmp = x[start:end]
        x_conflict_index = []
        for sub in tmp:
            index = y.index(sub)
            if not (index >= start and index < end):
                x_conflict_index.append(index)

        y_confict_index = []
        tmp = y[start:end]
        for sub in tmp:
            index = x.index(sub)
            if not (index >= start and index < end):
                y_confict_index.append(index)

        # 执行交叉
        x[start:end], y[start:end] = y[start:end], x[start:end]

        # 解决冲突
        for i in range(len(x_conflict_index)):
            x[y_confict_index[i]], y[x_conflict_index[i]] = y[x_conflict_index[i]], x[y_confict_index[i]]

        # 最终验证
        assert x[0] == 0 and x[-1] == len_ - 1
        assert y[0] == 0 and y[-1] == len_ - 1
        return x, y

    def ga_parent(self, scores, ga_choose_ratio):
        sort_index = np.argsort(-scores).copy()
        sort_index = sort_index[0:int(ga_choose_ratio * len(sort_index))]
        parents = []
        parents_score = []
        for index in sort_index:
            parents.append(self.fruits[index])
            parents_score.append(scores[index])
        return parents, parents_score

    def ga_choose(self, genes_score, genes_choose):
        sum_score = sum(genes_score)
        score_ratio = [sub * 1.0 / sum_score for sub in genes_score]
        rand1 = np.random.rand()  # 生成一个[0,1]的任意随机数
        rand2 = np.random.rand()
        for i, sub in enumerate(score_ratio):  # 同时列出数据和数据下标，一般用在for语句中
            if rand1 >= 0:
                rand1 -= sub
                if rand1 < 0:
                    index1 = i
            if rand2 >= 0:
                rand2 -= sub
                if rand2 < 0:
                    index2 = i
            if rand1 < 0 and rand2 < 0:
                break
        return list(genes_choose[index1]), list(genes_choose[index2])

    def ga_mutate_beifen(self, gene):
        path_list = [t for t in range(len(gene))]
        order = list(random.sample(path_list, 2))
        start, end = min(order), max(order)
        tmp = gene[start:end]
        # np.random.shuffle(tmp)
        tmp = tmp[::-1]  # 颠倒基因片段，变异
        gene[start:end] = tmp
        return list(gene)

    def ga_mutate(self, gene):
        if len(gene) <= 2:  # 只有起点和终点时不变异
            return gene
        # 只在中间部分(1~len_-2)选择变异区间
        start, end = sorted(random.sample(range(1, len(gene) - 1), 2))
        gene[start:end] = gene[start:end][::-1]  # 反转片段
        return gene

    def ga1(self):
        # 以前
        # 获得优质父代
        scores = self.compute_adp(self.fruits, self.length_total)
        # 选择部分优秀个体作为父代候选集合
        parents, parents_score = self.ga_parent(scores, self.ga_choose_ratio)
        tmp_best_one = parents[0]  # 精英保留策略
        tmp_best_score = parents_score[0]
        # 新的种群fruits
        fruits = parents.copy()
        # 生成新的种群
        while len(fruits) < self.num_total:
            # 轮盘赌方式对父代进行选择
            gene_x, gene_y = self.ga_choose(parents_score, parents)
            # 交叉
            gene_x_new, gene_y_new = self.ga_cross(gene_x, gene_y)
            # 变异
            if np.random.rand() < self.mutate_ratio:
                gene_x_new = self.ga_mutate(gene_x_new)
            if np.random.rand() < self.mutate_ratio:
                gene_y_new = self.ga_mutate(gene_y_new)
            x_adp = 1. / self.compute_pathlen(gene_x_new, self.dis_mat)
            y_adp = 1. / self.compute_pathlen(gene_y_new, self.dis_mat)
            # 将适应度高的放入种群中
            if x_adp > y_adp and (not gene_x_new in fruits):
                fruits.append(gene_x_new)
            elif x_adp <= y_adp and (not gene_y_new in fruits):
                fruits.append(gene_y_new)

        self.fruits = fruits
        return tmp_best_one, tmp_best_score

    def ga(self):
        # 现在
        scores = self.compute_adp(self.fruits, self.length_total)
        # 选择部分优秀个体作为父代候选集合
        parents, parents_score = self.ga_parent(scores, self.ga_choose_ratio)
        tmp_best_one = parents[0]  # 精英保留策略
        tmp_best_score = parents_score[0]
        # 新的种群
        new_fruits = parents.copy()

        # 修改后的种群生成逻辑
        while len(new_fruits) < self.num_total:
            # 轮盘赌选择
            gene_x, gene_y = self.ga_choose(parents_score, parents)
            # 交叉（增加交叉成功率检查）
            gene_x_new, gene_y_new = self.ga_cross(gene_x, gene_y)

            # 变异
            if np.random.rand() < self.mutate_ratio:
                gene_x_new = self.ga_mutate(gene_x_new)
            if np.random.rand() < self.mutate_ratio:
                gene_y_new = self.ga_mutate(gene_y_new)

            # 计算适应度（增加容错）
            x_adp = 1. / self.compute_pathlen(gene_x_new, self.dis_mat)
            y_adp = 1. / self.compute_pathlen(gene_y_new, self.dis_mat)
            # try:
            #     x_adp = 1. / max(0.0001, self.compute_pathlen(gene_x_new, self.dis_mat))
            #     y_adp = 1. / max(0.0001, self.compute_pathlen(gene_y_new, self.dis_mat))
            # except:
            #     continue

            # 放宽加入种群的条件
            if x_adp > y_adp:
                new_fruits.append(gene_x_new)
            else:
                new_fruits.append(gene_y_new)
            # if x_adp > y_adp and (not gene_x_new in new_fruits):
            #     new_fruits.append(gene_x_new)
            # elif x_adp <= y_adp and (not gene_y_new in new_fruits):
            #     new_fruits.append(gene_y_new)

            # 防止无限循环
            # if len(new_fruits) >= self.num_total * 2:
            #     break

        # 确保种群大小正确
        self.fruits = new_fruits
        return tmp_best_one, tmp_best_score

    def datedown_log(self, iteration, tmp10_best_one, txt_name):
        with open(txt_name, 'a+', encoding='utf-8') as f:
            content_wm = "{},{}\n".format(iteration, tmp10_best_one)
            f.write(content_wm)

    def run(self):
        try:
            print('run')
            BEST_LIST = None
            best_score = -math.inf
            self.best_record = []
            # WORK_DIR = r'E:\02Gitcode\gtus'
            self.txt_name = os.path.join(WORK_DIR, 'logs', self.uuid_value + '.txt')
            # print("————————————————————————改————————————————————")
            # self.txt_name = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs', self.uuid_value + '.txt')
            for i in range(1, self.iteration + 1):
                print(f"第{i}代")
                tmp_best_one, tmp_best_score = self.ga()
                self.iter_x.append(i)
                self.iter_y.append(1. / tmp_best_score)  # 当代最优个体适应度
                if tmp_best_score > best_score:
                    best_score = tmp_best_score
                    BEST_LIST = tmp_best_one
                self.best_record.append(1. / best_score)
                # if i % 10 == 0:
                #     self.datedown_log(i, self.best_record[i-1], self.txt_name)
            out_dict = dict(zip(tmp_best_one, self.location[BEST_LIST]))
            return self.location[BEST_LIST], 1. / best_score, out_dict
        except Exception as e:
            print(e)


def read_excel_to_data(path, start_point, end_point):
    f1 = xlrd.open_workbook(path)
    sheet = f1.sheet_by_index(0)  # 读序号为0的表
    rows = sheet.nrows
    data = [[] for x in range(rows + 1)]
    data[0] = start_point
    data[0].append(0)
    for i in range(1, rows):
        data[i] = sheet.row_values(i)[-2:]
        data[i].append(i)
    data[rows] = end_point
    data[rows].append(rows)
    return data


def _get_target_epsg(dataset):
    """从投影 shapefile 读取目标 EPSG 代码。"""
    prosrs, _ = getSRSPair(dataset)
    if prosrs is None:
        return 4549
    code = prosrs.GetAuthorityCode(None)
    if code:
        return int(code)
    return 4549


def lonlat2geo(dataset, lon, lat):
    """经纬度(EPSG:4490)转投影平面坐标。"""
    target_epsg = _get_target_epsg(dataset)
    transformer = Transformer.from_crs(4490, target_epsg, always_xy=True)
    x, y = transformer.transform(float(lon), float(lat))
    return [x, y]


def geo2lonlat(dataset, x, y):
    """投影平面坐标转经纬度(EPSG:4490)。"""
    target_epsg = _get_target_epsg(dataset)
    transformer = Transformer.from_crs(target_epsg, 4490, always_xy=True)
    lon, lat = transformer.transform(float(x), float(y))
    return [lat, lon]


def getSRSPair(path):  #
    driver = ogr.GetDriverByName('ESRI Shapefile')
    ds = driver.Open(path)
    layer0 = ds.GetLayerByIndex(0)
    return layer0.GetSpatialRef(), osr.SpatialReference()


def read_excel_xlsx_to_data(path, start_point, end_point):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    # 获取总行数
    rows = sheet.max_row
    data = [[] for x in range(rows + 1)]
    data[0] = start_point
    data[0].append(0)
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue
        else:
            data[i] = list(row)
            data[i].append(i)
    data[rows] = end_point
    data[rows].append(rows)
    return data


def generate_flight_candidates(
    path,
    progress_callback=None,
    grid_size=150,
    adaptive_grid=True,
):
    """生成规划区域内的候选航点，同时返回投影坐标和经纬度坐标。"""
    xls_path = os.path.join(os.path.dirname(path), "result.xlsx")
    prj_path = os.path.join(os.path.dirname(path), "input_prj.shp")
    get_finish_net_point(
        path,
        xls_path,
        prj_path,
        grid_size=grid_size,
        adaptive_grid=adaptive_grid,
        progress_callback=progress_callback,
    )

    workbook = load_workbook(filename=xls_path, read_only=True, data_only=True)
    sheet = workbook.active
    projected_points = []
    for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_index == 0:
            continue
        if row[0] is None or row[1] is None:
            continue
        projected_points.append([float(row[0]), float(row[1])])
    workbook.close()

    lonlat_points = []
    total_points = len(projected_points)
    for index, (x, y) in enumerate(projected_points, start=1):
        lat, lon = geo2lonlat(prj_path, x, y)
        lonlat_points.append([lon, lat])
        if progress_callback and (
            index == total_points or index % max(1, total_points // 20) == 0
        ):
            progress_callback(index, total_points, '正在转换航点坐标')
    return projected_points, lonlat_points, prj_path


def predict_flight_points(path, start_point, end_point, uuid_value):
    print(start_point, type(start_point))
    projected_points, _, prj_path = generate_flight_candidates(path)
    if start_point[0] < 180 and start_point[1] < 180:
        start_point = lonlat2geo(prj_path, start_point[0], start_point[1])
        end_point = lonlat2geo(prj_path, end_point[0], end_point[1])
    data = np.array([start_point] + projected_points + [end_point])

    num_city = data.shape[0]
    if num_city < 2:
        raise ValueError('规划区域未生成有效航点，请扩大绘制范围后重试')
    if num_city == 2:
        lonlat = geo2lonlat(prj_path, start_point[0], start_point[1])
        return [[lonlat[1], lonlat[0]], [lonlat[1], lonlat[0]]]

    Best, Best_path = math.inf, None
    model = GA(num_city=num_city, num_total=200, iteration=10, data=data.copy(), uuid_value=uuid_value)
    path_point, path_len, out_dict = model.run()  # out_dict为输出集合
    if path_len < Best:
        Best = path_len
        Best_path = path_point
    result_pts = []
    for xy in Best_path:
        lonlat = geo2lonlat(prj_path, xy[0], xy[1])
        result_pts.append([lonlat[1], lonlat[0]])
    result_ptss = np.array(result_pts)
    print(result_ptss)
    return result_pts


if __name__ == '__main__':
    path = r'C:\Users\Administrator\Desktop\test1\test1.shp'
    start_point = [119.81487686295512, 32.6051703664795]
    end_point = [119.81715174032428, 32.60542787782079]
    uuid_value = '6dc89a2a-b28e-4907-a335-d45aaaed094b'
    predict_flight_points(path, start_point, end_point, uuid_value)
